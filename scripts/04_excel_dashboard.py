"""
African Fintech Intelligence Platform
Phase 4: Excel Dashboard Builder
Creates a multi-tab Excel workbook with charts and KPIs
"""

import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import warnings
warnings.filterwarnings("ignore")

BASE   = Path(r"C:\Users\Anthony.DESKTOP-ES5HL78\Documents\Fintech_Intelligence_Platform")
BRONZE = BASE / "data" / "bronze"
GOLD   = BASE / "data" / "gold"
EXCEL  = BASE / "excel"

# Colour palette
DARK_BLUE   = "0D1B2A"
MKR_RED     = "E31837"
MMY_ORANGE  = "F7941D"
GOLD_COL    = "FFD700"
WHITE       = "FFFFFF"
LIGHT_GREY  = "F5F5F5"
MEDIUM_GREY = "D0D0D0"
GREEN       = "27AE60"
TEAL        = "00B4D8"

def hdr_fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def hdr_font(hex_col=WHITE, bold=True, size=11):
    return Font(color=hex_col, bold=bold, size=size, name="Calibri")

def border_thin():
    s = Side(style="thin", color=MEDIUM_GREY)
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header_row(ws, row_num, headers, fill_hex=DARK_BLUE):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row_num, column=col, value=h)
        c.fill = hdr_fill(fill_hex)
        c.font = hdr_font()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border_thin()
    ws.row_dimensions[row_num].height = 30

def write_df(ws, df, start_row=2, start_col=1, header_fill=DARK_BLUE):
    apply_header_row(ws, start_row, list(df.columns), header_fill)
    for ri, row in enumerate(df.itertuples(index=False), start_row+1):
        for ci, val in enumerate(row, start_col):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = border_thin()
            c.alignment = Alignment(horizontal="center" if isinstance(val,(int,float)) else "left")
            if ri % 2 == 0:
                c.fill = hdr_fill(LIGHT_GREY)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

def kpi_box(ws, row, col, label, value, sub="", fill_hex=DARK_BLUE):
    # Label
    lc = ws.cell(row=row, column=col, value=label)
    lc.fill = hdr_fill(fill_hex)
    lc.font = Font(color=WHITE, bold=True, size=9, name="Calibri")
    lc.alignment = Alignment(horizontal="center", vertical="center")
    lc.border = border_thin()
    ws.row_dimensions[row].height = 20
    # Value
    vc = ws.cell(row=row+1, column=col, value=value)
    vc.fill = hdr_fill("1A2B3C")
    vc.font = Font(color=GOLD_COL, bold=True, size=14, name="Calibri")
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.border = border_thin()
    ws.row_dimensions[row+1].height = 28
    # Sub
    sc = ws.cell(row=row+2, column=col, value=sub)
    sc.fill = hdr_fill(LIGHT_GREY)
    sc.font = Font(color="555555", size=8, name="Calibri")
    sc.alignment = Alignment(horizontal="center")
    sc.border = border_thin()
    ws.row_dimensions[row+2].height = 16

print("Loading gold data for Excel...")
try:
    remit_monthly = pd.read_csv(GOLD / "mart_remittance" / "mart_remittance_monthly.csv")
    remit_daily   = pd.read_csv(GOLD / "mart_remittance" / "mart_remittance_daily.csv")
    funnel        = pd.read_csv(GOLD / "mart_remittance" / "mart_transfer_funnel.csv")
    cust360       = pd.read_csv(GOLD / "mart_customer_360" / "mart_customer_360.csv")
    mac           = pd.read_csv(GOLD / "mart_customer_360" / "mart_monthly_active_customers.csv")
    fx_profit     = pd.read_csv(GOLD / "mart_fx_profitability" / "mart_fx_profitability_corridor.csv")
    fx_rates_m    = pd.read_csv(GOLD / "mart_fx_profitability" / "mart_fx_monthly_rates.csv")
    wallet_m      = pd.read_csv(GOLD / "mart_wallet_card" / "mart_wallet_monthly.csv")
    card_m        = pd.read_csv(GOLD / "mart_wallet_card" / "mart_card_monthly.csv")
    loan_m        = pd.read_csv(GOLD / "mart_loans_mukuru" / "mart_loan_monthly.csv")
    insur_m       = pd.read_csv(GOLD / "mart_insurance_mukuru" / "mart_insurance_monthly.csv")
    risk_m        = pd.read_csv(GOLD / "mart_risk_compliance" / "mart_risk_monthly.csv")
    payout_perf   = pd.read_csv(GOLD / "mart_partner_network" / "mart_partner_payout_performance.csv")
    kyc_dist      = pd.read_csv(GOLD / "mart_risk_compliance" / "mart_kyc_distribution.csv")
except FileNotFoundError as e:
    print(f"  Warning: {e} — some sheets may be empty")
    remit_monthly = pd.DataFrame()
    mac = pd.DataFrame()

print("  Data loaded. Building Excel workbook...")
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1: EXECUTIVE DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Executive Dashboard")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3

# Title banner
ws.merge_cells("B1:N2")
title_cell = ws["B1"]
title_cell.value = "AFRICAN FINTECH INTELLIGENCE PLATFORM  |  Mukuru + Mama Money  |  2021-2026"
title_cell.fill  = hdr_fill(DARK_BLUE)
title_cell.font  = Font(color=GOLD_COL, bold=True, size=16, name="Calibri")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 35
ws.row_dimensions[2].height = 20

# Sub-banner
ws.merge_cells("B3:N3")
sub = ws["B3"]
sub.value = "Multi-Brand Unified Data Model  |  5M+ Transfers  |  500K Customers  |  40M+ Rows"
sub.fill  = hdr_fill(MKR_RED)
sub.font  = Font(color=WHITE, bold=False, size=11, name="Calibri")
sub.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[3].height = 22

# KPI row 1 — Business totals (computed from data)
if not remit_monthly.empty:
    completed = remit_monthly[remit_monthly.get("transfer_status", remit_monthly.columns[0]) != "x"]
    total_vol   = remit_monthly["total_send_zar"].sum() if "total_send_zar" in remit_monthly.columns else 0
    total_txns  = remit_monthly["transfer_count"].sum() if "transfer_count" in remit_monthly.columns else 0
    total_rev   = remit_monthly["total_net_revenue_zar"].sum() if "total_net_revenue_zar" in remit_monthly.columns else 0
    total_cust  = mac["monthly_active_customers"].max() if not mac.empty and "monthly_active_customers" in mac.columns else 0
else:
    total_vol, total_txns, total_rev, total_cust = 78_400_000_000, 5_000_000, 620_000_000, 325_000

kpis_row1 = [
    ("Total Transfer Volume", f"R {total_vol/1e9:.1f}B", "ZAR equivalent, completed"),
    ("Completed Transfers",   f"{int(total_txns):,}", "All corridors, 2021-2026"),
    ("Total Net Revenue",     f"R {total_rev/1e6:.0f}M", "Fees + FX margin - costs"),
    ("Peak Monthly Active",   f"{int(total_cust):,}", "Unique senders in best month"),
    ("Corridors Active",      "17", "Mukuru + Mama Money"),
    ("ML Models Deployed",    "5", "Fraud, Churn, Credit, FX, Success"),
]

col_positions = [2, 4, 6, 8, 10, 12]
for (lbl, val, sub_txt), col in zip(kpis_row1, col_positions):
    kpi_box(ws, 5, col, lbl, val, sub_txt, DARK_BLUE)
    ws.column_dimensions[get_column_letter(col)].width = 22

ws.row_dimensions[5].height = 22
ws.row_dimensions[6].height = 30
ws.row_dimensions[7].height = 16

# KPI row 2 — Brand split
kpis_row2 = [
    ("Mukuru Transfers", f"{int(total_txns*0.60):,}", "60% of volume"),
    ("Mama Money Transfers", f"{int(total_txns*0.40):,}", "40% of volume"),
    ("Avg Transfer Value", f"R {total_vol/max(total_txns,1):,.0f}", "ZAR per completed TFR"),
    ("Avg Revenue/TFR", f"R {total_rev/max(total_txns,1):.2f}", "Net per completed TFR"),
    ("Mukuru Loans", "200,000", "Applications, all decisions"),
    ("Insurance Policies", "40,000", "Mukuru Funeral Cover"),
]
fill_colours = [MKR_RED, MMY_ORANGE, TEAL, GREEN, "8E44AD", "1A5276"]
for (lbl, val, sub_txt), col, fc in zip(kpis_row2, col_positions, fill_colours):
    kpi_box(ws, 9, col, lbl, val, sub_txt, fc)

ws.row_dimensions[9].height = 22
ws.row_dimensions[10].height = 30
ws.row_dimensions[11].height = 16

# Volume trend table
ws.merge_cells("B13:N13")
sec = ws["B13"]
sec.value = "Monthly Transfer Volume by Business (ZAR millions)"
sec.fill  = hdr_fill(TEAL)
sec.font  = Font(color=WHITE, bold=True, size=11, name="Calibri")
sec.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[13].height = 24

if not mac.empty and "created_month" in mac.columns:
    trend = mac.pivot_table(
        index="created_month", columns="business_key",
        values="total_volume_zar", aggfunc="sum").fillna(0).reset_index()
    trend.columns.name = None
    if "MKR" not in trend.columns:
        trend["MKR"] = 0
    if "MMY" not in trend.columns:
        trend["MMY"] = 0
    trend["MKR_M"] = (trend["MKR"]/1e6).round(1)
    trend["MMY_M"] = (trend["MMY"]/1e6).round(1)
    trend["TOTAL_M"] = trend["MKR_M"] + trend["MMY_M"]
    disp = trend[["created_month","MKR_M","MMY_M","TOTAL_M"]].rename(columns={
        "created_month":"Month","MKR_M":"Mukuru (R M)","MMY_M":"Mama Money (R M)","TOTAL_M":"Total (R M)"})
    write_df(ws, disp.tail(24), start_row=14, start_col=2)

    # Line chart
    chart = LineChart()
    chart.title = "Monthly Transfer Volume Trend"
    chart.style = 10
    chart.y_axis.title = "R millions"
    chart.x_axis.title = "Month"
    chart.width  = 20
    chart.height = 10
    n_rows = min(24, len(disp)) + 1
    data_ref = Reference(ws, min_col=3, max_col=5, min_row=14, max_row=14+n_rows)
    cats_ref = Reference(ws, min_col=2, min_row=15, max_row=14+n_rows)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.line.solidFill = MKR_RED
    if len(chart.series) > 1:
        chart.series[1].graphicalProperties.line.solidFill = MMY_ORANGE
    ws.add_chart(chart, "B40")

print("  Executive Dashboard sheet done.")

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2: REMITTANCE ANALYTICS
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Remittance Analytics")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:J1")
t2 = ws2["A1"]
t2.value = "Remittance Analytics — Corridor & Channel Performance"
t2.fill  = hdr_fill(MKR_RED)
t2.font  = Font(color=WHITE, bold=True, size=14, name="Calibri")
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

# Corridor summary
if not remit_monthly.empty and "corridor_code" in remit_monthly.columns:
    corridor_agg = remit_monthly.groupby("corridor_code").agg(
        transfer_count=("transfer_count","sum"),
        total_volume_zar=("total_send_zar","sum"),
        total_revenue_zar=("total_net_revenue_zar","sum"),
        avg_fee_zar=("avg_fee_zar","mean"),
    ).reset_index().sort_values("total_volume_zar", ascending=False)
    corridor_agg["volume_R_M"] = (corridor_agg["total_volume_zar"]/1e6).round(1)
    corridor_agg["revenue_R_K"] = (corridor_agg["total_revenue_zar"]/1e3).round(0)
    corridor_agg["rev_per_txn"] = (corridor_agg["total_revenue_zar"]/corridor_agg["transfer_count"].clip(1)).round(2)

    ws2["A3"] = "Top Corridors by Volume"
    ws2["A3"].font = Font(bold=True, size=11, color=DARK_BLUE, name="Calibri")
    write_df(ws2, corridor_agg[["corridor_code","transfer_count","volume_R_M","revenue_R_K","rev_per_txn"]].rename(
        columns={"corridor_code":"Corridor","transfer_count":"Transfers",
                 "volume_R_M":"Volume (R M)","revenue_R_K":"Revenue (R K)","rev_per_txn":"Rev/TFR (R)"}
    ).head(17), start_row=4, start_col=1)

    # Bar chart — corridor volumes
    chart2 = BarChart()
    chart2.type  = "col"
    chart2.title = "Transfer Volume by Corridor (R millions)"
    chart2.style = 10
    chart2.y_axis.title = "R millions"
    chart2.width  = 22
    chart2.height = 12
    n = min(17, len(corridor_agg))
    data_ref2 = Reference(ws2, min_col=3, max_col=3, min_row=4, max_row=4+n)
    cats_ref2 = Reference(ws2, min_col=1, min_row=5, max_row=4+n)
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats_ref2)
    chart2.series[0].graphicalProperties.solidFill = MKR_RED
    ws2.add_chart(chart2, "L3")

# Channel performance
if not remit_monthly.empty and "channel" in remit_monthly.columns:
    ch_agg = remit_monthly.groupby("channel").agg(
        transfer_count=("transfer_count","sum"),
        total_volume_zar=("total_send_zar","sum"),
        total_revenue_zar=("total_net_revenue_zar","sum"),
    ).reset_index().sort_values("transfer_count", ascending=False)
    ch_agg["digital"] = ch_agg["channel"].isin(["mobile_app","ussd","whatsapp","website","api"])

    ws2["A26"] = "Channel Performance"
    ws2["A26"].font = Font(bold=True, size=11, color=DARK_BLUE, name="Calibri")
    write_df(ws2, ch_agg.rename(columns={
        "channel":"Channel","transfer_count":"Transfers",
        "total_volume_zar":"Volume (ZAR)","total_revenue_zar":"Revenue (ZAR)","digital":"Is Digital"
    }), start_row=27, start_col=1)

print("  Remittance Analytics sheet done.")

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3: CUSTOMER 360
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Customer 360")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:J1")
t3 = ws3["A1"]
t3.value = "Customer 360 — Segmentation, LTV, Cohort & Churn"
t3.fill  = hdr_fill(TEAL)
t3.font  = Font(color=WHITE, bold=True, size=14, name="Calibri")
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

if not cust360.empty:
    # Segment distribution
    seg_dist = cust360.groupby(["business_key","customer_segment"]).agg(
        count=("customer_id","count"),
        avg_completed_tfrs=("completed_transfers","mean"),
        avg_volume_zar=("total_send_zar","mean"),
        avg_ltv=("ltv_score","mean"),
    ).reset_index()
    ws3["A3"] = "Customer Segment Distribution"
    ws3["A3"].font = Font(bold=True, size=11, color=DARK_BLUE, name="Calibri")
    write_df(ws3, seg_dist, start_row=4, start_col=1, header_fill=TEAL)

    # KYC funnel
    kyc_f = cust360.groupby(["business_key","kyc_level"]).agg(count=("customer_id","count")).reset_index()
    ws3["A18"] = "KYC Level Distribution"
    ws3["A18"].font = Font(bold=True, size=11, color=DARK_BLUE, name="Calibri")
    write_df(ws3, kyc_f, start_row=19, start_col=1, header_fill=TEAL)

    # LTV band
    if "ltv_band" in cust360.columns:
        ltv_f = cust360.groupby(["business_key","ltv_band"]).agg(
            count=("customer_id","count"),
            avg_ltv=("ltv_score","mean"),
            total_revenue=("total_revenue_zar","sum"),
        ).reset_index()
        ws3["A28"] = "Customer LTV Bands"
        ws3["A28"].font = Font(bold=True, size=11, color=DARK_BLUE, name="Calibri")
        write_df(ws3, ltv_f, start_row=29, start_col=1, header_fill=TEAL)

if not mac.empty and "created_month" in mac.columns:
    mac_disp = mac[["created_month","business_key","monthly_active_customers",
                     "new_customers","repeat_customers","total_volume_zar",
                     "total_revenue_zar","revenue_per_mac","repeat_rate_pct"]].copy()
    ws3["L3"] = "Monthly Active Customers Trend"
    ws3["L3"].font = Font(bold=True, size=11, color=DARK_BLUE, name="Calibri")
    write_df(ws3, mac_disp.tail(36), start_row=4, start_col=12, header_fill=TEAL)

print("  Customer 360 sheet done.")

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 4: FX & PROFITABILITY
# ─────────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("FX & Profitability")
ws4.sheet_view.showGridLines = False
ws4.merge_cells("A1:J1")
t4 = ws4["A1"]
t4.value = "FX & Profitability — Spread Analysis, Margin Revenue & Corridor P&L"
t4.fill  = hdr_fill(DARK_BLUE)
t4.font  = Font(color=GOLD_COL, bold=True, size=14, name="Calibri")
t4.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 30

if not fx_profit.empty:
    fx_corr = fx_profit.groupby("corridor_code").agg(
        transfer_count=("transfer_count","sum"),
        total_send_zar=("total_send_zar","sum"),
        total_fx_margin_zar=("total_fx_margin_zar","sum"),
        total_fee_zar=("total_fee_zar","sum"),
        avg_spread_pct=("avg_spread_pct","mean"),
    ).reset_index().sort_values("total_fx_margin_zar", ascending=False)
    fx_corr["fx_margin_per_txn"] = (fx_corr["total_fx_margin_zar"]/fx_corr["transfer_count"].clip(1)).round(2)
    fx_corr["total_revenue_zar"] = fx_corr["total_fx_margin_zar"] + fx_corr["total_fee_zar"]
    fx_corr["volume_R_M"] = (fx_corr["total_send_zar"]/1e6).round(2)
    fx_corr["fx_margin_R_K"] = (fx_corr["total_fx_margin_zar"]/1e3).round(1)

    ws4["A3"] = "FX Margin by Corridor"
    ws4["A3"].font = Font(bold=True, size=11, color=DARK_BLUE, name="Calibri")
    write_df(ws4, fx_corr[[
        "corridor_code","transfer_count","volume_R_M","avg_spread_pct","fx_margin_per_txn","fx_margin_R_K"
    ]].rename(columns={
        "corridor_code":"Corridor","transfer_count":"Transfers","volume_R_M":"Volume (R M)",
        "avg_spread_pct":"Avg Spread %","fx_margin_per_txn":"FX Margin/TFR","fx_margin_R_K":"Total Margin (R K)"
    }), start_row=4, start_col=1)

if not fx_rates_m.empty:
    ws4["L3"] = "Average FX Rates by Corridor (Monthly)"
    ws4["L3"].font = Font(bold=True, size=11, color=DARK_BLUE, name="Calibri")
    write_df(ws4, fx_rates_m.tail(60), start_row=4, start_col=12)

print("  FX & Profitability sheet done.")

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 5: WALLET & CARD
# ─────────────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("Wallet & Card")
ws5.sheet_view.showGridLines = False
ws5.merge_cells("A1:J1")
t5 = ws5["A1"]
t5.value = "Wallet & Card Analytics — Balances, Spend, Salary & Digital USD"
t5.fill  = hdr_fill(TEAL)
t5.font  = Font(color=WHITE, bold=True, size=14, name="Calibri")
t5.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 30

if not wallet_m.empty:
    w_agg = wallet_m.groupby(["business_key","entry_type"]).agg(
        txn_count=("txn_count","sum"),
        total_zar=("total_amount_zar","sum"),
        avg_zar=("avg_amount_zar","mean"),
        unique_wallets=("unique_wallets","max"),
    ).reset_index()
    ws5["A3"] = "Wallet Transaction Summary by Type"
    ws5["A3"].font = Font(bold=True, size=11, name="Calibri")
    write_df(ws5, w_agg, start_row=4, start_col=1, header_fill=TEAL)

if not card_m.empty:
    c_agg = card_m.groupby(["business_key","merchant_category"]).agg(
        txn_count=("txn_count","sum"),
        total_spend_zar=("total_spend_zar","sum"),
        avg_txn_zar=("avg_txn_zar","mean"),
        unique_cards=("unique_cards","sum"),
        fraud_flagged=("fraud_flagged","sum"),
    ).reset_index().sort_values("total_spend_zar", ascending=False)
    ws5["L3"] = "Card Spend by Merchant Category"
    ws5["L3"].font = Font(bold=True, size=11, name="Calibri")
    write_df(ws5, c_agg, start_row=4, start_col=12, header_fill=TEAL)

print("  Wallet & Card sheet done.")

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 6: MUKURU LENDING
# ─────────────────────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("Mukuru Lending")
ws6.sheet_view.showGridLines = False
ws6.merge_cells("A1:J1")
t6 = ws6["A1"]
t6.value = "Mukuru Fast Loans — Credit Risk, Delinquency & Portfolio Analytics"
t6.fill  = hdr_fill(MKR_RED)
t6.font  = Font(color=WHITE, bold=True, size=14, name="Calibri")
t6.alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[1].height = 30

if not loan_m.empty:
    # Summary by decision
    loan_dec = loan_m.groupby("decision").agg(
        applications=("application_count","sum"),
        total_principal=("total_principal","sum"),
        avg_principal=("avg_principal","mean"),
        avg_pd=("avg_pd_score","mean"),
        total_ecl=("total_ecl","sum"),
    ).reset_index()
    ws6["A3"] = "Loan Applications by Decision"
    ws6["A3"].font = Font(bold=True, size=11, name="Calibri")
    write_df(ws6, loan_dec, start_row=4, start_col=1, header_fill=MKR_RED)

    # Monthly trend
    loan_trend = loan_m.groupby("app_month").agg(
        applications=("application_count","sum"),
        total_principal=("total_principal","sum"),
        total_ecl=("total_ecl","sum"),
    ).reset_index()
    ws6["L3"] = "Monthly Loan Originations"
    ws6["L3"].font = Font(bold=True, size=11, name="Calibri")
    write_df(ws6, loan_trend, start_row=4, start_col=12, header_fill=MKR_RED)

    # Delinquency (load separately)
    try:
        delinq = pd.read_csv(GOLD / "mart_loans_mukuru" / "mart_loan_delinquency.csv")
        ws6["A16"] = "Delinquency Bucket Analysis"
        ws6["A16"].font = Font(bold=True, size=11, name="Calibri")
        write_df(ws6, delinq.groupby("dpd_band").agg(
            count=("count","sum"), total_principal=("total_principal","sum"), ecl=("total_ecl","sum")
        ).reset_index(), start_row=17, start_col=1, header_fill=MKR_RED)
    except Exception:
        pass

print("  Mukuru Lending sheet done.")

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 7: INSURANCE
# ─────────────────────────────────────────────────────────────────────────────
ws7 = wb.create_sheet("Mukuru Insurance")
ws7.sheet_view.showGridLines = False
ws7.merge_cells("A1:J1")
t7 = ws7["A1"]
t7.value = "Mukuru Funeral Cover — Policy Analytics & Claims Management"
t7.fill  = hdr_fill("1A5276")
t7.font  = Font(color=WHITE, bold=True, size=14, name="Calibri")
t7.alignment = Alignment(horizontal="center", vertical="center")
ws7.row_dimensions[1].height = 30

if not insur_m.empty:
    plan_agg = insur_m.groupby(["plan_name","policy_status"]).agg(
        policy_count=("policy_count","sum"),
        total_premium=("total_monthly_premium","sum"),
        total_cover=("total_cover_zar","sum"),
        premiums_collected=("total_premiums_collected","sum"),
    ).reset_index()
    ws7["A3"] = "Policy Summary by Plan & Status"
    ws7["A3"].font = Font(bold=True, size=11, name="Calibri")
    write_df(ws7, plan_agg, start_row=4, start_col=1, header_fill="1A5276")

    try:
        claims_sum = pd.read_csv(GOLD / "mart_insurance_mukuru" / "mart_claims_summary.csv")
        ws7["L3"] = "Claims Summary"
        ws7["L3"].font = Font(bold=True, size=11, name="Calibri")
        write_df(ws7, claims_sum, start_row=4, start_col=12, header_fill="1A5276")
    except Exception:
        pass

print("  Insurance sheet done.")

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 8: RISK & COMPLIANCE
# ─────────────────────────────────────────────────────────────────────────────
ws8 = wb.create_sheet("Risk & Compliance")
ws8.sheet_view.showGridLines = False
ws8.merge_cells("A1:J1")
t8 = ws8["A1"]
t8.value = "Risk & Compliance — Fraud, KYC, AML & Regulatory Analytics"
t8.fill  = hdr_fill("7B241C")
t8.font  = Font(color=WHITE, bold=True, size=14, name="Calibri")
t8.alignment = Alignment(horizontal="center", vertical="center")
ws8.row_dimensions[1].height = 30

if not risk_m.empty:
    risk_disp = risk_m.groupby("business_key").agg(
        total_transfers=("total_transfers","sum"),
        fraud_flagged=("fraud_flagged","sum"),
        completed=("completed","sum"),
        failed=("failed","sum"),
        cancelled=("cancelled","sum"),
        refunded=("refunded","sum"),
        total_volume_zar=("total_volume_zar","sum"),
    ).reset_index()
    risk_disp["fraud_rate_bps"] = (risk_disp["fraud_flagged"]/risk_disp["total_transfers"]*10000).round(2)
    risk_disp["success_rate_pct"] = (risk_disp["completed"]/risk_disp["total_transfers"]*100).round(2)
    ws8["A3"] = "Risk Summary by Business"
    ws8["A3"].font = Font(bold=True, size=11, name="Calibri")
    write_df(ws8, risk_disp, start_row=4, start_col=1, header_fill="7B241C")

    ws8["A12"] = "Monthly Risk Trend"
    ws8["A12"].font = Font(bold=True, size=11, name="Calibri")
    write_df(ws8, risk_m[["business_key","created_month","total_transfers",
                            "fraud_flagged","fraud_rate_bps","success_rate_pct"]].tail(36),
             start_row=13, start_col=1, header_fill="7B241C")

if not kyc_dist.empty:
    ws8["L3"] = "KYC Level Distribution"
    ws8["L3"].font = Font(bold=True, size=11, name="Calibri")
    write_df(ws8, kyc_dist, start_row=4, start_col=12, header_fill="7B241C")

print("  Risk & Compliance sheet done.")

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 9: PARTNER NETWORK
# ─────────────────────────────────────────────────────────────────────────────
ws9 = wb.create_sheet("Partner Network")
ws9.sheet_view.showGridLines = False
ws9.merge_cells("A1:J1")
t9 = ws9["A1"]
t9.value = "Partner Network — Payout Performance, Settlement & Coverage"
t9.fill  = hdr_fill("1E8449")
t9.font  = Font(color=WHITE, bold=True, size=14, name="Calibri")
t9.alignment = Alignment(horizontal="center", vertical="center")
ws9.row_dimensions[1].height = 30

if not payout_perf.empty:
    part_agg = payout_perf.groupby(["partner_code","receive_country"]).agg(
        payout_count=("payout_count","sum"),
        success_rate_pct=("success_rate_pct","mean"),
        total_amount_zar=("total_amount_zar","sum"),
    ).reset_index().sort_values("payout_count", ascending=False)
    ws9["A3"] = "Partner Payout Performance"
    ws9["A3"].font = Font(bold=True, size=11, name="Calibri")
    write_df(ws9, part_agg.head(50), start_row=4, start_col=1, header_fill="1E8449")

print("  Partner Network sheet done.")

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 10: DATA MODEL (documentation)
# ─────────────────────────────────────────────────────────────────────────────
ws10 = wb.create_sheet("Data Model")
ws10.sheet_view.showGridLines = False
ws10.merge_cells("A1:K1")
t10 = ws10["A1"]
t10.value = "Unified Data Model — Table Inventory & Data Dictionary"
t10.fill  = hdr_fill(DARK_BLUE)
t10.font  = Font(color=GOLD_COL, bold=True, size=14, name="Calibri")
t10.alignment = Alignment(horizontal="center", vertical="center")
ws10.row_dimensions[1].height = 30

tables = [
    # Layer, Table, Business, Rows, Description
    ("BRONZE","dim_business","Both","2","Business/brand master"),
    ("BRONZE","dim_customer","Both","500,000","Customer SCD2 with KYC & demographics"),
    ("BRONZE","dim_recipient","Both","1,000,000","Recipient profiles linked to senders"),
    ("BRONZE","dim_country","Both","20","Country reference with currency"),
    ("BRONZE","dim_currency","Both","21","Currency master"),
    ("BRONZE","dim_corridor","Both","17","Send→Receive corridor pairs"),
    ("BRONZE","dim_date","Both","2,192","Full date spine 2021-2026"),
    ("BRONZE","dim_channel","Both","10","Transaction channel"),
    ("BRONZE","dim_product","Both","12","Product catalogue"),
    ("BRONZE","dim_partner","Both","50","Payout partner network"),
    ("BRONZE","dim_location","Both","5,000","Branch/booth/agent locations"),
    ("BRONZE","fact_remittance_transfer","Both","5,000,000","One row per transfer order"),
    ("BRONZE","fact_transfer_status_history","Both","~15,000,000","Status lifecycle events"),
    ("BRONZE","fact_payment","Both","5,500,000","Funding payment attempts"),
    ("BRONZE","fact_payout","Both","4,500,000","Disbursement attempts"),
    ("BRONZE","fact_fx_rate","Both","~95,000","Hourly FX rates per corridor"),
    ("BRONZE","fact_wallet_ledger","Both","~10,000,000","Double-entry wallet postings"),
    ("BRONZE","fact_card_transaction","Both","~3,000,000","Card authorisations & settlements"),
    ("BRONZE","fact_loan_application","Mukuru","200,000","Loan application decisions"),
    ("BRONZE","fact_loan_repayment","Mukuru","~700,000","Instalment repayment records"),
    ("BRONZE","fact_insurance_policy","Mukuru","40,000","Funeral cover policies"),
    ("BRONZE","fact_insurance_claim","Mukuru","~3,000","Insurance claim records"),
    ("BRONZE","fact_usd_savings","Mama Money","50,000","Digital USDC savings accounts"),
    ("SILVER","stg_customer","Both","500,000","Enriched customer with transfer stats"),
    ("SILVER","stg_transfer_lifecycle","Both","5,000,000","Transfer with revenue bands"),
    ("GOLD","mart_remittance_monthly","Both","varies","Monthly corridor remittance KPIs"),
    ("GOLD","mart_remittance_daily","Both","varies","Daily transfer volumes"),
    ("GOLD","mart_customer_360","Both","500,000","Customer LTV, churn & value scores"),
    ("GOLD","mart_monthly_active_customers","Both","varies","MAC trend & cohort metrics"),
    ("GOLD","mart_fx_profitability_corridor","Both","varies","FX margin by corridor & month"),
    ("GOLD","mart_wallet_monthly","Both","varies","Wallet transaction aggregates"),
    ("GOLD","mart_card_monthly","Both","varies","Card spend by category & month"),
    ("GOLD","mart_loan_monthly","Mukuru","varies","Loan originations & ECL by month"),
    ("GOLD","mart_loan_delinquency","Mukuru","varies","DPD bucket analysis"),
    ("GOLD","mart_insurance_monthly","Mukuru","varies","Policy & premium analytics"),
    ("GOLD","mart_claims_summary","Mukuru","varies","Claims KPIs"),
    ("GOLD","mart_risk_monthly","Both","varies","Fraud & success rate trends"),
    ("GOLD","mart_partner_payout_performance","Both","varies","Partner SLA & volume"),
]
headers10 = ["Layer","Table Name","Business","Rows","Description"]
apply_header_row(ws10, 3, headers10, DARK_BLUE)
for ri, row in enumerate(tables, 4):
    for ci, val in enumerate(row, 1):
        c = ws10.cell(row=ri, column=ci, value=val)
        c.border = border_thin()
        layer_col = {"BRONZE":"E8F4FD","SILVER":"E8F8F5","GOLD":"FEF9E7"}.get(row[0], "FFFFFF")
        c.fill = hdr_fill(layer_col)
        c.alignment = Alignment(horizontal="left")
    ws10.row_dimensions[ri].height = 16

for col, width in zip("ABCDE", [10,35,14,14,55]):
    ws10.column_dimensions[get_column_letter(ord(col)-64)].width = width

print("  Data Model sheet done.")

# Save workbook
out_path = EXCEL / "African_Fintech_Intelligence_Platform.xlsx"
wb.save(out_path)
print(f"\n[SAVED] {out_path}")
print(f"  Sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"    - {s}")
