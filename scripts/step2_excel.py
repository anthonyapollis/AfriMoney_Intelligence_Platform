"""
AfriMoney Intelligence Platform — Excel Dashboard
10 tabs with charts, KPIs, ML results, corridor data.
All numbers aligned with ebook and map.
"""
import json, datetime
from pathlib import Path
import numpy as np

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable,"-m","pip","install","openpyxl","-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.utils import get_column_letter

BASE    = Path(r"C:\Users\Anthony.DESKTOP-ES5HL78\Documents\Fintech_Intelligence_Platform")
XL_DIR  = BASE / "excel"
XL_DIR.mkdir(exist_ok=True)
ML_FILE = BASE / "ml_models" / "ml_results.json"

with open(ML_FILE) as f:
    ml = json.load(f)

# ── PALETTE ────────────────────────────────────────────────────
G  = "00C97A"   # AfriMoney green
GO = "F5A623"   # gold
CO = "FF6B6B"   # coral
SK = "4ECDC4"   # sky
NV = "0D2137"   # navy
W  = "FFFFFF"
LG = "F4F6F8"   # light grey
MG = "DDE1E7"   # mid grey
MKR= "00C97A"
MMY= "F5A623"

def fill(hex_): return PatternFill("solid", fgColor=hex_)
def bold(size=11,color=W,italic=False):
    return Font(bold=True,size=size,color=color,italic=italic,name="Segoe UI")
def reg(size=10,color="1A2332"):
    return Font(size=size,color=color,name="Segoe UI")
def centre(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
def left():   return Alignment(horizontal="left",vertical="center",wrap_text=True)
def border():
    s=Side(border_style="thin",color=MG)
    return Border(left=s,right=s,top=s,bottom=s)

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

def make_sheet(name, tab_color):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab_color
    return ws

def header_row(ws, row, cols, texts, bg=NV, fg=W, height=32):
    ws.row_dimensions[row].height = height
    for c,t in zip(cols,texts):
        cell = ws.cell(row=row,column=c,value=t)
        cell.font      = bold(11,fg)
        cell.fill      = fill(bg)
        cell.alignment = centre()
        cell.border    = border()

def kpi_box(ws, row, col, label, value, bg=G, fg=W, w=18, h=40):
    ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height   = 16
    ws.row_dimensions[row+1].height = 28
    ws.row_dimensions[row+2].height = 14
    lc = ws.cell(row=row,  column=col,value=label)
    lc.font=Font(size=9,color=bg,name="Segoe UI",bold=True)
    lc.alignment=centre()
    vc = ws.cell(row=row+1,column=col,value=value)
    vc.font=bold(18,fg)
    vc.fill=fill(bg)
    vc.alignment=centre()
    bc = ws.cell(row=row+2,column=col,value="")
    bc.fill=fill(bg)

# ══════════════════════════════════════════════════════════════
# TAB 1 — EXECUTIVE DASHBOARD
# ══════════════════════════════════════════════════════════════
ws1 = make_sheet("Executive Dashboard", G)
ws1.column_dimensions["A"].width = 2
ws1.sheet_view.showGridLines = False

# Title banner
ws1.merge_cells("B1:N2")
tc = ws1["B1"]
tc.value = "AfriMoney Intelligence Platform — Executive Dashboard"
tc.font  = bold(18,W)
tc.fill  = fill(NV)
tc.alignment = centre()
ws1.row_dimensions[1].height = 20
ws1.row_dimensions[2].height = 20

ws1.merge_cells("B3:N3")
sub = ws1["B3"]
sub.value = f"Snowflake · dbt · Snowpark ML  |  Reporting Period: 2025–2026  |  Generated: {datetime.date.today()}"
sub.font  = Font(size=10,color="AAAAAA",italic=True,name="Segoe UI")
sub.fill  = fill(NV)
sub.alignment = centre()
ws1.row_dimensions[3].height = 18

# KPI row
kpis = [
    ("Annual Volume",     "R 78B",  G),
    ("Transfer Orders",   "5M+",    GO),
    ("Active Corridors",  "17",     SK),
    ("Customers",         "500K",   "A569BD"),
    ("Success Rate",      "78.2%",  G),
    ("Fraud Rate",        "3.2 bps",CO),
    ("KYC Completion",    "82%",    GO),
    ("Net Revenue",       "R 18B",  SK),
]
for i,(lbl,val,bg) in enumerate(kpis):
    col = 2 + i*2
    ws1.merge_cells(start_row=5,start_column=col,end_row=5,end_column=col+1)
    ws1.merge_cells(start_row=6,start_column=col,end_row=6,end_column=col+1)
    ws1.merge_cells(start_row=7,start_column=col,end_row=7,end_column=col+1)
    lc=ws1.cell(row=5,column=col,value=lbl)
    lc.font=Font(size=9,bold=True,color=bg,name="Segoe UI")
    lc.fill=fill("F8FAFB"); lc.alignment=centre()
    vc=ws1.cell(row=6,column=col,value=val)
    vc.font=bold(20,bg); vc.fill=fill("F8FAFB"); vc.alignment=centre()
    ws1.cell(row=7,column=col).fill=fill("F8FAFB")
    for r in [5,6,7]:
        ws1.cell(row=r,column=col).border=border()
    ws1.row_dimensions[5].height=14
    ws1.row_dimensions[6].height=36
    ws1.row_dimensions[7].height=10

# Monthly trend data
months = ["Jan-26","Feb-26","Mar-26","Apr-26","May-26","Jun-26"]
mkr_vol  = [8.2,8.9,9.4,10.1,10.8,11.4]
mmy_vol  = [4.1,4.5,4.8,5.2,5.6,5.9]
mkr_rev  = [1.89,2.05,2.16,2.32,2.48,2.62]
mmy_rev  = [0.82,0.90,0.96,1.04,1.12,1.18]
success  = [76.1,76.8,77.4,78.0,78.6,79.2]
fraud_bps= [4.1,3.8,3.5,3.2,3.0,2.8]

ws1.row_dimensions[9].height = 16
ws1.cell(row=9,column=2).value = "Monthly Trend Data"
ws1.cell(row=9,column=2).font  = bold(12,NV)
ws1.cell(row=9,column=2).fill  = fill(LG)

header_row(ws1,10,[2,3,4,5,6,7,8,9],
           ["Month","MKR Vol (R B)","MMY Vol (R B)","Total Vol (R B)","MKR Rev (R B)","MMY Rev (R B)","Success Rate %","Fraud Rate bps"],
           bg="0D2137")

for i,(m,mv,sv,mrv,srv,sr,fr) in enumerate(zip(months,mkr_vol,mmy_vol,mkr_rev,mmy_rev,success,fraud_bps)):
    r=11+i
    ws1.row_dimensions[r].height=20
    vals=[m,mv,sv,round(mv+sv,1),mrv,srv,sr,fr]
    bg_row = LG if i%2==0 else W
    for j,v in enumerate(vals):
        c=ws1.cell(row=r,column=2+j,value=v)
        c.fill=fill(bg_row); c.border=border(); c.alignment=centre()
        c.font=reg()
        if j==6: c.font=Font(size=10,color=G,bold=True,name="Segoe UI")
        if j==7: c.font=Font(size=10,color=CO,bold=True,name="Segoe UI")

# Bar chart — volume by month
chart = BarChart()
chart.type="col"; chart.grouping="clustered"
chart.title="Monthly Transfer Volume (R Billion)"
chart.y_axis.title="Volume (R B)"; chart.x_axis.title="Month"
chart.width=18; chart.height=12
chart.style=10
data = Reference(ws1,min_col=3,max_col=5,min_row=10,max_row=16)
cats = Reference(ws1,min_col=2,min_row=11,max_row=16)
chart.add_data(data,titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.solidFill=MKR
chart.series[1].graphicalProperties.solidFill=MMY
chart.series[2].graphicalProperties.solidFill=SK
ws1.add_chart(chart,"B18")

# Line chart — success rate
lc2=LineChart()
lc2.title="Transfer Success Rate & Fraud (bps)"
lc2.width=18; lc2.height=10; lc2.style=10
d1=Reference(ws1,min_col=8,max_col=8,min_row=10,max_row=16)
d2=Reference(ws1,min_col=9,max_col=9,min_row=10,max_row=16)
lc2.add_data(d1,titles_from_data=True)
lc2.add_data(d2,titles_from_data=True)
lc2.set_categories(cats)
lc2.series[0].graphicalProperties.line.solidFill=G
lc2.series[0].graphicalProperties.line.width=25000
lc2.series[1].graphicalProperties.line.solidFill=CO
lc2.series[1].graphicalProperties.line.width=25000
ws1.add_chart(lc2,"J18")

# ══════════════════════════════════════════════════════════════
# TAB 2 — CORRIDOR ANALYSIS
# ══════════════════════════════════════════════════════════════
ws2 = make_sheet("Corridor Analysis", GO)
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 2

ws2.merge_cells("B1:L2")
h=ws2["B1"]; h.value="Corridor Performance Analysis — All 17 Routes"
h.font=bold(16,W); h.fill=fill(NV); h.alignment=centre()
ws2.row_dimensions[1].height=22; ws2.row_dimensions[2].height=22

corridors = [
    ("ZA-ZW","Zimbabwe","High-Volume Core","MKR+MMY",18.4,820000,3.2,78.2,5.41,1240),
    ("ZA-MZ","Mozambique","High-Volume Core","MKR+MMY",12.1,540000,2.8,76.8,5.38,1150),
    ("ZA-ZM","Zambia","High-Volume Core","MKR",8.7,390000,2.1,79.4,5.39,1320),
    ("ZA-MW","Malawi","High-Volume Core","MKR",6.2,280000,3.5,77.1,5.52,980),
    ("ZA-LS","Lesotho","Premium","MKR+MMY",4.8,195000,1.8,82.3,6.21,1680),
    ("ZA-SZ","Eswatini","Premium","MMY",3.1,128000,1.4,84.1,6.05,1510),
    ("ZA-BW","Botswana","Premium","MKR+MMY",5.4,210000,2.0,81.7,5.88,1720),
    ("ZA-NA","Namibia","Premium","MKR",3.9,142000,1.9,80.5,5.95,1590),
    ("ZA-MU","Mauritius","Premium","MMY",2.2,64000,0.9,87.4,6.44,2140),
    ("ZA-TZ","Tanzania","Growth","MKR+MMY",4.1,168000,4.1,74.3,5.15,1080),
    ("ZA-KE","Kenya","Growth","MMY",3.6,148000,3.8,75.6,5.22,1140),
    ("ZA-GH","Ghana","Growth","MKR+MMY",2.9,118000,4.6,73.1,5.08,1050),
    ("ZA-NG","Nigeria","Growth","MMY",3.4,138000,6.1,69.8,4.85,1180),
    ("ZA-MG","Madagascar","Growth","MKR",1.8,72000,5.2,71.2,4.92,890),
    ("ZA-AO","Angola","Emerging","MKR",1.4,52000,6.8,68.4,4.71,920),
    ("ZA-CD","DR Congo","Emerging","MKR",1.1,44000,8.2,64.2,4.55,870),
    ("ZA-ET","Ethiopia","Emerging","MKR",0.9,36000,7.1,66.8,4.68,810),
]

cluster_colors = {"High-Volume Core":CO,"Premium":GO,"Growth":SK,"Emerging":G}

header_row(ws2,4,[2,3,4,5,6,7,8,9,10,11],
           ["Corridor","Country","ML Cluster","Business","Vol (R B)","Transfers","Fraud bps","Success %","FX Spread %","Avg Send R"])

for i,(code,country,cluster,biz,vol,cnt,fraud,suc,fx,avg) in enumerate(corridors):
    r=5+i; ws2.row_dimensions[r].height=20
    bg_r=LG if i%2==0 else W
    vals=[code,country,cluster,biz,vol,cnt,fraud,suc,fx,avg]
    for j,v in enumerate(vals):
        c=ws2.cell(row=r,column=2+j,value=v)
        c.fill=fill(bg_r); c.border=border(); c.alignment=centre(); c.font=reg()
    # colour cluster cell
    clc=ws2.cell(row=r,column=4)
    clc.font=Font(size=10,bold=True,color=cluster_colors.get(cluster,NV),name="Segoe UI")
    # colour success
    sc=ws2.cell(row=r,column=9)
    col_s=G if suc>=80 else (GO if suc>=70 else CO)
    sc.font=Font(size=10,bold=True,color=col_s,name="Segoe UI")
    # colour fraud
    fc=ws2.cell(row=r,column=8)
    col_f=G if fraud<3 else (GO if fraud<6 else CO)
    fc.font=Font(size=10,bold=True,color=col_f,name="Segoe UI")

# Bar chart — corridor volumes, colour by cluster group
corr_bar_colors=["FF6B6B","FF6B6B","FF6B6B","FF6B6B",  # high-volume = coral
                 "F5A623","F5A623","F5A623","F5A623","F5A623",  # premium = gold
                 "4ECDC4","4ECDC4","4ECDC4","4ECDC4","4ECDC4",  # growth = sky
                 "00C97A","00C97A","00C97A"]  # emerging = green
bc2=BarChart(); bc2.type="bar"; bc2.grouping="clustered"
bc2.title="Corridor Volume (R Billion) — colour = ML cluster"; bc2.width=20; bc2.height=16; bc2.style=10
bc2.legend=None
d=Reference(ws2,min_col=6,max_col=6,min_row=4,max_row=21)
c2=Reference(ws2,min_col=3,min_row=5,max_row=21)
bc2.add_data(d,titles_from_data=True); bc2.set_categories(c2)
for idx,cl in enumerate(corr_bar_colors):
    pt=DataPoint(idx=idx); pt.graphicalProperties.solidFill=cl
    bc2.series[0].data_points.append(pt)
ws2.add_chart(bc2,"B24")

# ══════════════════════════════════════════════════════════════
# TAB 3 — ML MODEL RESULTS
# ══════════════════════════════════════════════════════════════
ws3 = make_sheet("ML Model Results", "A569BD")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 2

ws3.merge_cells("B1:M2")
h=ws3["B1"]; h.value="Machine Learning Model Results — AfriMoney Intelligence Platform"
h.font=bold(16,W); h.fill=fill("6C3483"); h.alignment=centre()
ws3.row_dimensions[1].height=22; ws3.row_dimensions[2].height=22

# Model summary table
header_row(ws3,4,[2,3,4,5,6,7,8,9,10],
           ["Model","Algorithm","Train N","Test N","AUC-ROC","Avg Precision","CV Mean","CV Std","Inference Mode"],
           bg="6C3483")

model_rows = [
    ("Fraud Detection","GradientBoostingClassifier",ml["fraud"]["n_train"],ml["fraud"]["n_test"],
     ml["fraud"]["auc"],ml["fraud"]["ap"],ml["fraud"]["cv_mean"],ml["fraud"]["cv_std"],"Real-time UDF"),
    ("Customer Churn","RandomForestClassifier",ml["churn"]["n_train"],ml["churn"]["n_test"],
     ml["churn"]["auc"],ml["churn"]["ap"],ml["churn"]["cv_mean"],ml["churn"]["cv_std"],"Weekly batch"),
    ("Credit Risk PD","GradientBoostingClassifier",ml["credit"]["n_train"],ml["credit"]["n_test"],
     ml["credit"]["auc"],ml["credit"]["ap"],ml["credit"]["cv_mean"],ml["credit"]["cv_std"],"At origination"),
]
auc_colors=[G,SK,GO]
for i,(name,algo,ntr,nte,auc,ap,cvm,cvs,inf) in enumerate(model_rows):
    r=5+i; ws3.row_dimensions[r].height=24
    bg_r=LG if i%2==0 else W
    vals=[name,algo,f"{ntr:,}",f"{nte:,}",auc,ap,cvm,cvs,inf]
    for j,v in enumerate(vals):
        c=ws3.cell(row=r,column=2+j,value=v)
        c.fill=fill(bg_r); c.border=border(); c.alignment=centre(); c.font=reg()
    # highlight AUC
    ac=ws3.cell(row=r,column=6)
    ac.font=Font(size=12,bold=True,color=auc_colors[i],name="Segoe UI")

# Feature importance tables for each model
fi_labels = {"fraud":"Fraud Detection","churn":"Customer Churn","credit":"Credit Risk PD"}
fi_colors  = {"fraud":G,"churn":SK,"credit":GO}
fi_cols    = {"fraud":2,"churn":6,"credit":10}

ws3.row_dimensions[9].height=18
ws3.cell(row=9,column=2).value="Feature Importance by Model (Top 8)";
ws3.cell(row=9,column=2).font=bold(12,NV); ws3.cell(row=9,column=2).fill=fill(LG)

for key,label in fi_labels.items():
    sc=fi_cols[key]; cl=fi_colors[key]
    ws3.cell(row=10,column=sc).value=label
    ws3.cell(row=10,column=sc).font=bold(11,cl)
    ws3.cell(row=10,column=sc).fill=fill(LG)
    ws3.cell(row=11,column=sc).value="Feature"
    ws3.cell(row=11,column=sc).font=bold(10,NV)
    ws3.cell(row=11,column=sc).fill=fill(MG)
    ws3.cell(row=11,column=sc+1).value="Importance"
    ws3.cell(row=11,column=sc+1).font=bold(10,NV)
    ws3.cell(row=11,column=sc+1).fill=fill(MG)
    ws3.column_dimensions[get_column_letter(sc)].width=22
    ws3.column_dimensions[get_column_letter(sc+1)].width=12
    for j,(feat,imp) in enumerate(ml[key]["fi"].items()):
        r=12+j; ws3.row_dimensions[r].height=18
        bg_r=LG if j%2==0 else W
        fc=ws3.cell(row=r,column=sc,value=feat.replace("_"," ").title())
        fc.fill=fill(bg_r); fc.border=border(); fc.font=reg(); fc.alignment=left()
        ic=ws3.cell(row=r,column=sc+1,value=round(imp,4))
        ic.fill=fill(bg_r); ic.border=border(); ic.font=reg(); ic.alignment=centre()

# Bar charts for feature importance — each bar gets its own colour
fi_bar_palettes = {
    "fraud":  ["FF6B6B","FF8E8E","FFB3B3","F5A623","FFD700","00C97A","4ECDC4","A569BD"],
    "churn":  ["4ECDC4","00C97A","F5A623","A569BD","FF6B6B","3498DB","1ABC9C","E67E22"],
    "credit": ["F5A623","FF6B6B","00C97A","A569BD","4ECDC4","3498DB","E74C3C","95A5A6"],
}
for key,col_off,cl in [("fraud",2,G),("churn",6,SK),("credit",10,GO)]:
    n_feats = len(ml[key]["fi"])
    bc=BarChart(); bc.type="bar"; bc.grouping="clustered"
    bc.title=fi_labels[key]+" — Feature Importance"
    bc.width=14; bc.height=11; bc.style=10
    bc.legend=None  # colours tell the story; no need for redundant legend
    d=Reference(ws3,min_col=col_off+1,max_col=col_off+1,min_row=11,max_row=11+n_feats)
    c=Reference(ws3,min_col=col_off,min_row=12,max_row=11+n_feats)
    bc.add_data(d,titles_from_data=True); bc.set_categories(c)
    palette = fi_bar_palettes[key]
    for idx in range(n_feats):
        pt=DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = palette[idx % len(palette)]
        bc.series[0].data_points.append(pt)
    anchor_col={"fraud":"B","churn":"F","credit":"J"}[key]
    ws3.add_chart(bc,f"{anchor_col}22")

# ══════════════════════════════════════════════════════════════
# TAB 4 — CUSTOMER 360
# ══════════════════════════════════════════════════════════════
ws4 = make_sheet("Customer 360", SK)
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 2

ws4.merge_cells("B1:K2")
h=ws4["B1"]; h.value="Customer 360 — Segmentation & Lifetime Value"
h.font=bold(16,W); h.fill=fill("1A9E98"); h.alignment=centre()
ws4.row_dimensions[1].height=22; ws4.row_dimensions[2].height=22

segments=[
    ("Champion","LTV ≥ 2000","MKR+MMY","LEVEL_3",2.8,4200,95,8200,1.2,42000),
    ("High Value","LTV 500–2000","MKR+MMY","LEVEL_2+",18.4,2800,88,3100,8.1,28500),
    ("Medium Value","LTV 100–500","MKR","LEVEL_2",44.1,1200,72,680,22.8,12400),
    ("Low Value","LTV 1–100","MKR+MMY","LEVEL_1+",26.3,480,58,140,11.8,4200),
    ("Zero Value","LTV = 0","MKR+MMY","LEVEL_0+",8.4,0,31,0,1.8,0),
]

header_row(ws4,4,[2,3,4,5,6,7,8,9,10,11],
           ["LTV Band","Definition","Business","KYC Level","Cust % of 500K","Avg LTV Score","Avg Engagement","Lifetime Rev R","Churn Rate %","Wallet Bal R"],
           bg="1A9E98")

ltv_colors=["FFD700","F5A623","A569BD","4ECDC4","BDC3C7"]
for i,(band,defn,biz,kyc,pct,ltv,eng,rev,churn,wal) in enumerate(segments):
    r=5+i; ws4.row_dimensions[r].height=22
    bg_r=LG if i%2==0 else W
    # pct stored as numeric (col 6) so the pie chart can read it; display with % format
    vals=[band,defn,biz,kyc,pct,ltv,f"{eng}%",f"R {rev:,}",f"{churn}%",f"R {wal:,}"]
    for j,v in enumerate(vals):
        c=ws4.cell(row=r,column=2+j,value=v)
        c.fill=fill(bg_r); c.border=border(); c.alignment=centre(); c.font=reg()
    ws4.cell(row=r,column=2).font=Font(size=11,bold=True,color=ltv_colors[i],name="Segoe UI")
    # format pct cell as percentage display
    ws4.cell(row=r,column=6).number_format='0.0"%"'

# Pie chart — customer distribution (uses numeric pct column, rows 5-9, no header row)
pc=PieChart(); pc.title="Customer Distribution by LTV Band"
pc.width=14; pc.height=12; pc.style=10
d=Reference(ws4,min_col=6,min_row=5,max_row=9)   # numeric data only, no header
c=Reference(ws4,min_col=2,min_row=5,max_row=9)    # labels: LTV Band names
pc.add_data(d,titles_from_data=False); pc.set_categories(c)
pie_colors=["FFD700","F5A623","A569BD","4ECDC4","BDC3C7"]
for j,cl in enumerate(pie_colors):
    pt=DataPoint(idx=j); pt.graphicalProperties.solidFill=cl
    pc.series[0].data_points.append(pt)
ws4.add_chart(pc,"B12")

# KYC distribution
ws4.cell(row=12,column=9).value="KYC Level Distribution"
ws4.cell(row=12,column=9).font=bold(11,NV)
kyc_data=[("LEVEL_0","Unverified","5%",CO),("LEVEL_1","Basic","20%",GO),
          ("LEVEL_2","Verified","50%",G),("LEVEL_3","Premium","25%",SK)]
for i,(lv,nm,pct,cl) in enumerate(kyc_data):
    r=13+i; ws4.row_dimensions[r].height=20
    ws4.cell(row=r,column=9,value=lv).font=Font(size=10,bold=True,color=cl,name="Segoe UI")
    ws4.cell(row=r,column=10,value=nm).font=reg()
    ws4.cell(row=r,column=11,value=pct).font=Font(size=10,bold=True,color=cl,name="Segoe UI")
    for c in [9,10,11]:
        ws4.cell(row=r,column=c).fill=fill(LG if i%2==0 else W)
        ws4.cell(row=r,column=c).border=border()
        ws4.cell(row=r,column=c).alignment=centre()

# ══════════════════════════════════════════════════════════════
# TAB 5 — FX & PROFITABILITY
# ══════════════════════════════════════════════════════════════
ws5 = make_sheet("FX & Profitability", GO)
ws5.sheet_view.showGridLines = False
ws5.column_dimensions["A"].width = 2

ws5.merge_cells("B1:J2")
h=ws5["B1"]; h.value="FX Rates & Corridor Profitability"
h.font=bold(16,W); h.fill=fill("D4890A"); h.alignment=centre()
ws5.row_dimensions[1].height=22; ws5.row_dimensions[2].height=22

fx_data=[
    ("ZA-ZW","USD/ZWG","5.41%","4.82%","0.59%","R 4.82B",SK),
    ("ZA-MZ","ZAR/MZN","5.38%","4.78%","0.60%","R 3.21B",G),
    ("ZA-LS","ZAR/LSL","6.21%","5.55%","0.66%","R 1.48B",GO),
    ("ZA-MU","ZAR/MUR","6.44%","5.76%","0.68%","R 0.71B",CO),
    ("ZA-BW","ZAR/BWP","5.88%","5.24%","0.64%","R 1.68B",SK),
    ("ZA-SZ","ZAR/SZL","6.05%","5.41%","0.64%","R 0.99B",G),
    ("ZA-ZM","ZAR/ZMW","5.39%","4.80%","0.59%","R 2.69B",GO),
    ("ZA-NG","NGN/ZAR","4.85%","4.19%","0.66%","R 0.87B",CO),
    ("ZA-KE","ZAR/KES","5.22%","4.61%","0.61%","R 1.10B",SK),
    ("ZA-TZ","ZAR/TZS","5.15%","4.54%","0.61%","R 1.27B",G),
]

header_row(ws5,4,[2,3,4,5,6,7],
           ["Corridor","Currency Pair","FX Spread %","Fee Share %","Net FX Margin %","FX Revenue"],
           bg="D4890A")

for i,(code,pair,spread,fee,net,rev,cl) in enumerate(fx_data):
    r=5+i; ws5.row_dimensions[r].height=20
    bg_r=LG if i%2==0 else W
    for j,v in enumerate([code,pair,spread,fee,net,rev]):
        c=ws5.cell(row=r,column=2+j,value=v)
        c.fill=fill(bg_r); c.border=border(); c.alignment=centre()
        c.font=Font(size=10,bold=(j==0),color=cl if j==0 else "1A2332",name="Segoe UI")

# ══════════════════════════════════════════════════════════════
# TAB 6 — RISK & COMPLIANCE
# ══════════════════════════════════════════════════════════════
ws6 = make_sheet("Risk & Compliance", CO)
ws6.sheet_view.showGridLines = False
ws6.column_dimensions["A"].width = 2

ws6.merge_cells("B1:J2")
h=ws6["B1"]; h.value="Risk & Compliance Dashboard — Monthly KPIs"
h.font=bold(16,W); h.fill=fill("C0392B"); h.alignment=centre()
ws6.row_dimensions[1].height=22; ws6.row_dimensions[2].height=22

header_row(ws6,4,[2,3,4,5,6,7,8,9],
           ["Month","Fraud Rate bps","Fraud RAG","KYC Complete %","KYC RAG","Success Rate %","Ops RAG","Cancelled %"],
           bg="C0392B")

risk_data=[
    ("Jan-26",4.1,"🟡",79.2,"🟡",76.1,"🟡",8.2),
    ("Feb-26",3.8,"🟡",79.8,"🟡",76.8,"🟡",7.9),
    ("Mar-26",3.5,"🟡",80.5,"🟢",77.4,"🟢",7.6),
    ("Apr-26",3.2,"🟢",81.2,"🟢",78.0,"🟢",7.3),
    ("May-26",3.0,"🟢",81.8,"🟢",78.6,"🟢",7.0),
    ("Jun-26",2.8,"🟢",82.4,"🟢",79.2,"🟢",6.7),
]
for i,(m,fr,fr_rag,kyc,kyc_rag,sr,ops_rag,canc) in enumerate(risk_data):
    r=5+i; ws6.row_dimensions[r].height=22
    bg_r=LG if i%2==0 else W
    for j,v in enumerate([m,fr,fr_rag,kyc,kyc_rag,sr,ops_rag,canc]):
        c=ws6.cell(row=r,column=2+j,value=v)
        c.fill=fill(bg_r); c.border=border(); c.alignment=centre(); c.font=reg()
    cl_fr=G if fr<3 else(GO if fr<5 else CO)
    cl_sr=G if sr>=80 else(GO if sr>=70 else CO)
    ws6.cell(row=r,column=3).font=Font(size=10,bold=True,color=cl_fr,name="Segoe UI")
    ws6.cell(row=r,column=7).font=Font(size=10,bold=True,color=cl_sr,name="Segoe UI")

# ══════════════════════════════════════════════════════════════
# TAB 7 — LENDING (MUKURU)
# ══════════════════════════════════════════════════════════════
ws7 = make_sheet("AfriMoney Lending", NV)
ws7.sheet_view.showGridLines = False
ws7.column_dimensions["A"].width = 2

ws7.merge_cells("B1:J2")
h=ws7["B1"]; h.value="AfriMoney Lending Suite — Origination, IFRS 9 & Credit Risk"
h.font=bold(16,W); h.fill=fill(NV); h.alignment=centre()
ws7.row_dimensions[1].height=22; ws7.row_dimensions[2].height=22

# Product split note
ws7.merge_cells("B3:J3")
note=ws7["B3"]
note.value="MKR: Mukuru Fast Loan (cash-backed micro-loans)  |  MMY: Mama Money Payday Advance (pilot, R200–R1000)  |  Both under AfriMoney brand"
note.font=Font(size=9,italic=True,color="AAAAAA",name="Segoe UI")
note.fill=fill(NV); note.alignment=centre()
ws7.row_dimensions[3].height=16

loan_kpis=[("Total Applications","215K",G),("Approved","65%",GO),
           ("Default Rate","11.2%",CO),("Avg Loan Size","R 3,890",SK),
           ("IFRS 9 ECL Rate","4.8%",GO),("AUC Credit Model",str(ml["credit"]["auc"]),G)]
for i,(lbl,val,cl) in enumerate(loan_kpis):
    c=3+i*2
    ws7.merge_cells(start_row=4,start_column=c,end_row=4,end_column=c+1)
    ws7.merge_cells(start_row=5,start_column=c,end_row=5,end_column=c+1)
    lc=ws7.cell(row=4,column=c,value=lbl); lc.font=Font(size=9,bold=True,color=cl,name="Segoe UI")
    lc.fill=fill(LG); lc.alignment=centre()
    vc=ws7.cell(row=5,column=c,value=val); vc.font=bold(16,cl)
    vc.fill=fill(LG); vc.alignment=centre()
    ws7.row_dimensions[4].height=14; ws7.row_dimensions[5].height=32

header_row(ws7,7,[2,3,4,5,6,7,8,9],
           ["Month","Applications","Approved","Approval %","Disbursed R M","Defaults","Default %","ECL R M"],
           bg=NV)
loan_monthly=[
    ("Jan-26",3200,2100,65.6,8.82,96,4.6,0.42),
    ("Feb-26",3400,2210,65.0,9.28,101,4.6,0.44),
    ("Mar-26",3650,2380,65.2,9.98,109,4.6,0.47),
    ("Apr-26",3500,2275,65.0,9.56,104,4.6,0.45),
    ("May-26",3800,2470,65.0,10.37,113,4.6,0.49),
    ("Jun-26",4100,2665,65.0,11.19,122,4.6,0.53),
]
for i,row_data in enumerate(loan_monthly):
    r=8+i; ws7.row_dimensions[r].height=20
    bg_r=LG if i%2==0 else W
    for j,v in enumerate(row_data):
        c=ws7.cell(row=r,column=2+j,value=v)
        c.fill=fill(bg_r); c.border=border(); c.alignment=centre(); c.font=reg()

# MKR vs MMY product split
ws7.row_dimensions[15].height=18
ws7.cell(row=15,column=2).value="Business Split — Lending Products"
ws7.cell(row=15,column=2).font=bold(11,NV); ws7.cell(row=15,column=2).fill=fill(LG)

header_row(ws7,16,[2,3,4,5,6,7],
           ["Business","Product","Applications","Avg Loan R","Default %","Market"],
           bg=NV)
split_data=[
    ("Mukuru (MKR)","Mukuru Fast Loan",200000,4200,11.2,"13 corridors — app + USSD"),
    ("Mama Money (MMY)","MMY Payday Advance (pilot)",15000,850,9.8,"ZA-LS, ZA-SZ, ZA-BW only"),
]
for i,(biz,prod,apps,avg,dpct,mkt) in enumerate(split_data):
    r=17+i; ws7.row_dimensions[r].height=22
    bg_r=LG if i%2==0 else W
    cl=MKR if i==0 else MMY
    for j,v in enumerate([biz,prod,f"{apps:,}",f"R {avg:,}",f"{dpct}%",mkt]):
        c=ws7.cell(row=r,column=2+j,value=v)
        c.fill=fill(bg_r); c.border=border(); c.alignment=centre()
        c.font=Font(size=10,bold=(j==0),color=cl if j==0 else "1A2332",name="Segoe UI")

# ══════════════════════════════════════════════════════════════
# TAB 8 — WALLET & CARD
# ══════════════════════════════════════════════════════════════
ws8 = make_sheet("Wallet & Card", G)
ws8.sheet_view.showGridLines = False
ws8.column_dimensions["A"].width = 2

ws8.merge_cells("B1:J2")
h=ws8["B1"]; h.value="Wallet & Card Product Analytics"
h.font=bold(16,W); h.fill=fill("0A3D2E"); h.alignment=centre()
ws8.row_dimensions[1].height=22; ws8.row_dimensions[2].height=22

header_row(ws8,4,[2,3,4,5,6,7,8],
           ["Month","Wallet Active","Cash In R M","Cash Out R M","Card Txns","Card Spend R M","Card Approval %"],
           bg="0A3D2E")

wallet_data=[
    ("Jan-26",182000,245,218,310000,42.1,87.2),
    ("Feb-26",188000,258,231,324000,44.4,87.8),
    ("Mar-26",196000,271,244,341000,46.8,88.1),
    ("Apr-26",204000,285,256,358000,49.2,88.4),
    ("May-26",213000,299,268,376000,51.7,88.8),
    ("Jun-26",222000,314,281,394000,54.2,89.1),
]
for i,row_data in enumerate(wallet_data):
    r=5+i; ws8.row_dimensions[r].height=20
    bg_r=LG if i%2==0 else W
    for j,v in enumerate(row_data):
        c=ws8.cell(row=r,column=2+j,value=v)
        c.fill=fill(bg_r); c.border=border(); c.alignment=centre(); c.font=reg()

# ══════════════════════════════════════════════════════════════
# TAB 9 — DATA MODEL
# ══════════════════════════════════════════════════════════════
ws9 = make_sheet("Data Model", NV)
ws9.sheet_view.showGridLines = False
ws9.column_dimensions["A"].width = 2

ws9.merge_cells("B1:K2")
h=ws9["B1"]; h.value="AfriMoney — Snowflake Data Model (Bronze→Silver→Gold)"
h.font=bold(16,W); h.fill=fill(NV); h.alignment=centre()
ws9.row_dimensions[1].height=22; ws9.row_dimensions[2].height=22

tables=[
    ("BRONZE","dim_customer","500,000","SCD2","customer_sk","customer_id, kyc_level, risk_band, status",G),
    ("BRONZE","dim_recipient","1,000,000","Dimension","recipient_id","name_hash, country, payout_method",G),
    ("BRONZE","fact_remittance_transfer","5,000,000","Fact","transfer_id","send_amt_zar, corridor, channel, status",G),
    ("BRONZE","fact_wallet_ledger","~10,000,000","Fact (double-entry)","ledger_id","customer_id, direction, amount, type",G),
    ("BRONZE","fact_card_transaction","~3,000,000","Fact","txn_id","amount, status, merchant, fraud_flag",G),
    ("BRONZE","fact_loan_application","200,000","Fact","loan_id","principal, decision, dpd, pd_score",G),
    ("BRONZE","fact_fx_rate","~95,000","Fact","rate_id","corridor, interbank, applied, spread_pct",G),
    ("BRONZE","fact_insurance_policy","40,000","Fact","policy_id","product, premium, status",G),
    ("BRONZE","fact_usd_savings","50,000","Fact","account_id","balance_zar, usdc_balance, status",G),
    ("SILVER","stg_transfers","view","dbt view","—","Cleaned transfers + derived fields",SK),
    ("SILVER","int_transfer_profitability","~5M","dbt table","—","Transfers + FX join + margin tiers",SK),
    ("SILVER","int_customer_transfer_stats","~500K","dbt table","—","Aggregated customer behaviour",SK),
    ("GOLD","mart_remittance","~50K","dbt table","—","Monthly corridor KPIs + revenue",GO),
    ("GOLD","mart_customer_360","~500K","dbt table","—","One row per customer + LTV/engagement",GO),
    ("GOLD","mart_fx_profitability","~200","dbt table","—","Corridor FX margin by month",GO),
    ("GOLD","mart_risk_compliance","~12","dbt table","—","Monthly risk RAG status KPIs",GO),
    ("GOLD","mart_loans_mukuru","~1K","dbt table","—","Loan origination + ECL + repayment",GO),
    ("ML_DB","FRAUD_FEATURES","~200K","Feature table","—","15 fraud features refreshed daily",CO),
    ("ML_DB","CHURN_FEATURES","~500K","Feature table","—","21 churn features refreshed weekly",CO),
]

header_row(ws9,4,[2,3,4,5,6,7],
           ["Schema","Table Name","Row Count","Type","Key Column","Key Fields"],
           bg=NV)

layer_colors={"BRONZE":CO,"SILVER":SK,"GOLD":GO,"ML_DB":"A569BD"}
for i,(schema,tbl,rows,typ,key,fields,_) in enumerate(tables):
    r=5+i; ws9.row_dimensions[r].height=20
    bg_r=LG if i%2==0 else W
    cl=layer_colors.get(schema,NV)
    for j,v in enumerate([schema,tbl,rows,typ,key,fields]):
        c=ws9.cell(row=r,column=2+j,value=v)
        c.fill=fill(bg_r); c.border=border()
        c.font=Font(size=10,bold=(j==0 or j==1),color=cl if j==0 else "1A2332",name="Segoe UI")
        c.alignment=left() if j in [1,5] else centre()
    ws9.column_dimensions[get_column_letter(3)].width=28
    ws9.column_dimensions[get_column_letter(7)].width=40

# ══════════════════════════════════════════════════════════════
# TAB 10 — ABOUT
# ══════════════════════════════════════════════════════════════
ws10= make_sheet("About", G)
ws10.sheet_view.showGridLines = False

ws10.merge_cells("B2:H3")
h=ws10["B2"]; h.value="AfriMoney Intelligence Platform"
h.font=bold(22,W); h.fill=fill(NV); h.alignment=centre()

about=[
    ("Platform","AfriMoney Intelligence Platform — African Remittance Analytics"),
    ("Author","Anthony Apollis"),
    ("Date",str(datetime.date.today())),
    ("Tech Stack","Snowflake · dbt · Snowpark ML · Python · Power BI"),
    ("Architecture","Bronze → Silver → Gold (Medallion)"),
    ("Data Volume","40M+ rows across 20 table types"),
    ("ML Models","Fraud Detection (AUC "+str(ml["fraud"]["auc"])+") · Churn ("+str(ml["churn"]["auc"])+") · Credit PD ("+str(ml["credit"]["auc"])+")"),
    ("dbt Models","14 models · 28 tests passing"),
    ("Corridors","17 active SA-Africa remittance corridors"),
    ("Disclaimer","Synthetic data only. No real customer data used."),
]
for i,(k,v) in enumerate(about):
    r=5+i; ws10.row_dimensions[r].height=22
    kc=ws10.cell(row=r,column=2,value=k)
    kc.font=bold(11,G); kc.fill=fill(LG); kc.border=border(); kc.alignment=left()
    ws10.column_dimensions["B"].width=18
    vc=ws10.cell(row=r,column=3,value=v)
    vc.font=reg(11); vc.fill=fill(W); vc.border=border(); vc.alignment=left()
    ws10.column_dimensions["C"].width=55
    ws10.merge_cells(start_row=r,start_column=3,end_row=r,end_column=8)

# ── SAVE ─────────────────────────────────────────────────────
out = XL_DIR / "AfriMoney_Dashboard.xlsx"
wb.save(out)
print(f"[SAVED] {out}")
print(f"  Sheets: {len(wb.sheetnames)}")
print(f"  Sheets: {', '.join(wb.sheetnames)}")
